import functools
import io
from datetime import datetime
from itertools import product
from pathlib import Path
import re
from typing import Union, Tuple, Callable, List

import openpyxl
import pandas as pd
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, ExcelWriter, ExcelFile

from fram.generelle_hjelpemoduler.konstanter import (
    SKIPSTYPER,
    LENGDEGRUPPER,
    TRAFIKK_COLS,
    FRAM_DIRECTORY,
)


def les_inn_tankested(
    filbane: Union[pd.ExcelFile, Path, str], logger: Callable = print
) -> Tuple[str, str]:

    definisjon_df = pd.read_excel(
        filbane, sheet_name="Definisjoner", usecols=[1], skiprows=3, nrows=2
    )

    tankested = definisjon_df.iloc[0].values[0].lower()

    if tankested.lower() not in ["sør", "nord"]:
        logger(
            'Fant ikke tankersted i "Definisjoner"-arket i inputfilen (celle 5B). \nGyldige verdier er "sør" eller "nord"\nSetter tankersted lik "NORD" (for Trondheim)\n'
        )
        tankested = "nord"

    return tankested

@functools.lru_cache(3)
def les_inn_bruttoliste_pakker_skip_lengder(filbane):
    """Leser inn en bruttoliste med alle mulige kombinasjoner av ruter, skipstyper og lengdegrupper til bruk i trafikkberegningene"""
    godkjent_rutekolonne = re.compile("Rute\d+")
    wide_rute_df = (
        pd
        .read_excel(filbane, sheet_name="Ruteoversikt")
        .loc[:, lambda df: [col for col in df.columns if col in ["Strekning", "Tiltaksomraade", "Analyseomraade", "Tiltakspakke"] or godkjent_rutekolonne.match(col)]]
        .dropna(how="all")
        .astype({
        "Strekning": str,
        "Tiltaksomraade": int,
        "Tiltakspakke": int,
    })
    )
    trafikkgrupper = pd.DataFrame(
        list(product(SKIPSTYPER, LENGDEGRUPPER)),
        columns=["Skipstype", "Lengdegruppe"],
    )
    return (
        pd.wide_to_long(
            wide_rute_df,
            stubnames="Rute",
            i=["Strekning", "Tiltaksomraade", "Tiltakspakke", "Analyseomraade"],
            j="drop",
        )
        .reset_index()
        .drop("drop", axis=1)
        .dropna(subset=["Rute"])
        .assign(dummy=1)
        .merge(right=trafikkgrupper.assign(dummy=1), on="dummy", how="outer")
        .drop("dummy", axis=1)[TRAFIKK_COLS]
        .sort_values(by=TRAFIKK_COLS)
        .assign(Tiltaksomraade=lambda x: x.Tiltaksomraade.astype(int))
        .set_index(TRAFIKK_COLS)
    )


def _parse_strekning(strekning: Union[str, Path]):
    """
    Hjelpemetode for å lese og forstå strekningsnavnet som angis som input til FRAM. Dette kan være en
    streng eller en Path. Metoden sjekker at filen finnes, og at det er en excel-fil

    Args:
        strekning: Filbanen til en gyldig Excel-fil med input for en strekning

    Returns:
        Path: En vasket og kontrollert filbane til input-filen.

    """
    if strekning is None:
        raise KeyError("Kan ikke kalle modellen uten angitt strekning. Fikk None")
    elif isinstance(strekning, str):
        try:
            strekning = Path(strekning)
        except (ValueError, AssertionError):
            raise KeyError(f"Kan ikke se at '{strekning}' er en fil")

    if not (
        isinstance(strekning, Path) and strekning.is_file()
    ):  # Feil hvis dette ikke er en fil
        raise KeyError(f"Kan ikke se at '{strekning}' er en fil")

    if not strekning.suffix == ".xlsx":
        raise KeyError(f"'strekning' må være en 'xlsx'-fil, ikke {strekning.suffix}")
    return strekning

def _fra_excel(filbane: Union[ExcelFile, Path, str], ark: str, **kwargs):
    """
    Hjelpemetode for å lese inn et spesifikt ark i en excelbok. Alle ark som
    leses inn må ha kolonner trafikk_cols (Strekning, Tiltaksomrrade, Tiltakspakke
    Analyseomraade,Rute,Skipstype,Lengdegruppe )

    Args:
        - filbane: filbane til excelboken som ønskes leses inn.
        - ark: navn på arket som ønskes lest inn fra excelboken.

    Returns:
        En dataframe med innlest data fra et bestemt ark i en bestemt excelbok som
    har indeks lik trafikk_cols.
    """
    df = pd.read_excel(filbane, sheet_name=ark, **kwargs)
    # håndterer duplikater, bør erstattes dersom mangle_dupe_cols implementeres
    for i in range(1, 10):
        df.columns = df.columns.map(
            lambda x: (isinstance(x, str) and x.endswith(f".{i}") and x[:-2]) or x
        )

    assert df.columns.nunique() == len(df.columns)
    df = df.set_index(TRAFIKK_COLS)
    return df


def _fyll_ut_fra_alle(df, kolonne, fyllverdier):
    """
    Hjelpemetode som tar alle verdier med "Alle" i en bestemt "kolonne"
    og erstatter dem med nye rader med hver av fyllverdiene. Et eksempel på dette
    er dersom man har spesifisert en bestemt seilingstid for "Alle" skipstyper. I
    inputarket kan man da spesifisere "Alle", og ved å bruke denne hjelpemetoden
    vil man kunne spre tidsverdiene for "Alle" på skipstypene man ønsker, som da
    vil være "fyllverdier".

    Args:
        - df: en dataframe man ønsker å fikse på
        - kolonne: kolonnen som har verdien "Alle"
        - fyllverdier: hvilke fyllverdier man ønsker å duplisere verdiene som "Alle"
          er spesifisert for.

    Returns:
        En dataframe som har dupliserte rader for alle fyllverdier.
    """

    def erstatt_verdier_med_serier(df, kolonne, fyllverdier_ved_alle):
        """Hjelpefunksjon som bytter ut 'Alle' med en serie med fyllverdier"""
        df = df.copy()
        df[kolonne] = df[kolonne].map(
            lambda x: pd.Series(fyllverdier_ved_alle) if x == "Alle" else pd.Series(x)
        )
        return df

    df2 = (
        df.copy()
        .pipe(
            erstatt_verdier_med_serier,
            kolonne=kolonne,
            fyllverdier_ved_alle=fyllverdier,
        )
        .apply(lambda x: pd.Series(x[kolonne]), axis=1)
        .stack()
        .rename(kolonne)
        .reset_index(level=1, drop=True)
        .to_frame()
        .join(df.drop(kolonne, axis=1))
        .reset_index(drop=True)
    )
    return df2


def _lag_excel_forside(filepath, strekning, tiltakspakke, version, input_filbane, log):
    """
    Lager forside med informasjon om kjøring til exceloutput.

    Args:
        filepath: filbane og navn på excelbok forsiden skal lages i
        strekning: strekningen kjøringen av FRAM er gjort for
        tiltakspakke: tiltakspakken kjøringen av FRAM er gjort for
        version: versjon av modellen kjøringen av FRAM er gjort for
        input_filbane: inputboken som ble brukt til kjøringen
        log:

    Returns:

    """
    KYB_BLAA = "1F588E"  # Kystverket-blå

    # Sjekker om det finnes en Excel-bok der. Hvis ikke, lager den en ny.
    #try:
    wb = load_workbook(filepath)
    #except:
    #wb = Workbook()
    # Sletter forsidearket hvis det finnes, her spares det altså ikke på noe informasjon fra tidligere
    #try:
    #    existing_sheet = wb["Forside"]
    #    wb.remove(existing_sheet)
    #except:
    #    pass
    # Oppretter så arket på nytt
    sheet = wb.create_sheet("Forside", index=0)

    # Setter fargen på Excel-fanen og alle cellene
    sheet.sheet_properties.tabColor = KYB_BLAA
    for row in sheet["A1:AF100"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFFFFF")

    try:  # Skriver Kystverkets logo til forsiden
        try:
            import requests
        except ImportError:
            print(
                "Advarsel: Forsøkte å lage forside til FRAM, men trenger pakken requests for å gjøre det. Denne er ikke installert."
            )
        logo_path = "https://www.kystverket.no/contentassets/aced4444422a42d58e0e25d8cbbaa5b1/logofiler/kystverket-fullfarger.png"
        image = Image.open(io.BytesIO(requests.get(logo_path).content))
        img = openpyxl.drawing.image.Image(image)
        img.width /= 8
        img.height /= 8
        sheet.add_image(img, "L8")
    except:
        pass

    # Skriver header og setter stor blå font på den
    sheet["A1"] = f"Output fra FRAM -  {strekning} tiltakspakke {tiltakspakke}"
    sheet["A1"].font = Font(size=36, bold=True, color=KYB_BLAA)

    # Skrver info om FRAM-versjonen, strekningen, tiltakspakken, tidspunktet og input-filen.
    sheet[
        "A2"
    ] = f"Dette er output fra {version}, Kystverkets beregningsverktøy for samfunnsøkonomiske virkninger. Output gjelder tiltakspakke {tiltakspakke} på {strekning}"

    sheet[
        "A4"
    ] = f"Modellen ble kjørt {datetime.today().strftime('%Y-%m-%d kl. %H:%M')} fra inputfilen {input_filbane}"

    # Sjekker om den er på en git-gren og skriver i så fall den inputen.
    try:
        git_mappe = FRAM_DIRECTORY.parent / ".git"
        current_branch = (git_mappe / "HEAD").read_text().split("/")[-1].strip()
        latest_commit = (git_mappe / "refs" / "heads" / current_branch).read_text()
        git_string = f"Modellen ble kjørt fra git-grenen {current_branch}. Siste git-commit er {latest_commit}."
    except:
        git_string = "Fant ingen git-info. Koden kjøres antakelig utenfor git."

    sheet["A5"] = git_string

    # Skriver loggen fra FRAM-objektet til excel-filen.
    sheet["A10"] = "Under ligger loggen fra kjøringen"
    _skriv_log_excel_forside(filepath=wb, log=log)
    if filepath is not None:
        wb.active = 0
        # Har slitt med at ark nummer to også er valgt når man åpner Excel. Setter det derfor her til uvalgt eksplisitt
        wb["Resultater"].sheet_view.tabSelected = False
        wb.save(filepath)


def _skriv_log_excel_forside(filepath, log, start_row=11, start_col=1):
    if isinstance(filepath, Workbook):
        wb = filepath
    else:
        wb = load_workbook(filepath)
    try:
        sheet = wb["Forside"]
    except:
        raise KeyError(
            f"Du forsøkte å skrive FRAM-loggen til Excel-forsiden, men modellen fant ikke et ark 'Forside' i Excel-boken {filepath}. Har du kalt denne funksjonen manuelt? Det er ikke meningen."
        )
    else:
        assert isinstance(sheet, Worksheet)
    for idx, content in enumerate(log):
        sheet.cell(row=start_row + idx, column=start_col, value=content)
    try:
        wb.save()
    except:
        pass


def les_inn_ventetidssituasjoner_fra_excel(
    filbane: Union[Path, str], tiltakspakke: int
):
    ventetid_ref = []
    ventetid_tiltak = []
    wb = load_workbook(filbane, data_only=True)
    sheetnames = wb.sheetnames
    for sheet in sheetnames:
        if (
            ("ventetid" in sheet)
            & ("ref" in sheet)
            & (str(wb[sheet]["B8"].value) == str(tiltakspakke))
        ):  # Navnet på input-sheet for referansebanen må inneholde "ventetid" og "ref"
            ventetid_ref.append([sheet, wb[sheet]["B10"].value])
        elif (
            ("ventetid" in sheet)
            & ("tiltak" in sheet)
            & (str(wb[sheet]["B8"].value) == str(tiltakspakke))
        ):  # Navnet på input-sheet for tiltaksbanen må inneholde "ventetid" og "tiltak"
            ventetid_tiltak.append([sheet, wb[sheet]["B10"].value])
    ventetid_ref = pd.DataFrame(ventetid_ref, columns=["ark_ref", "Rute"])
    ventetid_tiltak = pd.DataFrame(ventetid_tiltak, columns=["ark_tiltak", "Rute"])

    ventetid_input = ventetid_tiltak.merge(
        ventetid_ref, on="Rute", indicator=True, how="outer"
    )
    # Gjør sjekk av at alle par er komplette og skriver feilmelding hvis de ikke er det
    if ventetid_input.loc[ventetid_input["_merge"] == "right_only"].count()[1] > 0:
        print(
            "Feilmelding: "
            + str(
                ventetid_input.loc[ventetid_input["_merge"] == "right_only"].count()[1]
            )
            + " av rutene har spesifisert ventetidsinput i referansebanen, men ikke i tiltaksbanen - inkluderes ikke i beregningene"
        )
    if ventetid_input.loc[ventetid_input["_merge"] == "left_only"].count()[1] > 0:
        print(
            "Feilmelding: "
            + str(
                ventetid_input.loc[ventetid_input["_merge"] == "left_only"].count()[1]
            )
            + " av rutene har spesifisert ventetidsinput i tiltaksbanen, men ikke i referansebanen - inkluderes ikke i beregningene"
        )
    ventetid_input = ventetid_input.loc[
        ventetid_input["_merge"] == "both", ["Rute", "ark_ref", "ark_tiltak"]
    ]
    return ventetid_input

def angi_kolonnenavn(df: DataFrame, kolonnenavn: List[str]) -> DataFrame:
    """ Setter listen med kolonnenavn som kolonnenavnene i df"""
    df = df.copy()
    antall_kolonner, antall_navn = df.shape[1], len(kolonnenavn)
    assert antall_kolonner == antall_navn, f"Prøver å sette kolonnenavn på en datframe med {antall_kolonner} kolonner og {antall_navn} navn"
    df.columns = kolonnenavn
    return df


def vask_kolonnenavn_for_exceltull(df):
    columns = df.columns.tolist()
    new_cols = []
    for col in columns:
        for i in range(10):
            if f".{i}" in str(col):
                col = col.rstrip(f".{i}")
        new_cols.append(col)
    df.columns = new_cols
    return df
